import os
import sqlite3

import openpyxl
from aiogram import types

from system.dispatcher import dp


# Функция для создания файла Excel с данными заказов
def create_excel_file(orders):
    """Запись данных в Excel"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # Заголовки столбцов
    sheet['A1'] = 'Имя'
    sheet['B1'] = 'Номер телефона'
    sheet['C1'] = 'Вопрос'
    sheet['D1'] = 'Дата вопроса'
    # Заполнение данными заказов
    for index, order in enumerate(orders, start=2):
        sheet.cell(row=index, column=1).value = order[0]  # Имя
        sheet.cell(row=index, column=2).value = order[1]  # Номер телефона
        sheet.cell(row=index, column=3).value = order[2]  # Вопрос
        sheet.cell(row=index, column=4).value = order[3]  # Дата вопроса

    return workbook


# Обработчик команды для выгрузки данных в Excel
@dp.message_handler(commands=['export_users'])
async def export_data(message: types.Message):
    """Обработчик команды /export_users"""
    if message.from_user.id not in [535185511, 6224881647]:
        await message.reply('У вас нет доступа к этой команде.')
        return
    # Подключение к базе данных SQLite
    conn = sqlite3.connect('setting/orders.db')
    cursor = conn.cursor()
    # Получение данных из базы данных
    cursor.execute("SELECT * FROM user_question")
    orders = cursor.fetchall()
    # Создание файла Excel
    workbook = create_excel_file(orders)
    # Сохранение файла
    filename = 'question.xlsx'
    workbook.save(filename)
    # Отправка файла пользователю
    with open(filename, 'rb') as file:
        await message.answer_document(file)
    # Удаление файла
    os.remove(filename)


def admin_handler():
    """Регистрируем handlers для выгрузки данных в Excel"""
    dp.register_message_handler(export_data)  # Обработчик команды /start, он же пост приветствия
